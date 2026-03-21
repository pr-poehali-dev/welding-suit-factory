"""
Отправка заявки на email менеджера + сохранение контакта в БД.
POST / — принимает данные формы или корзины калькулятора.
К письму прикрепляется Excel-файл с данными заявки.
"""
import json
import os
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SMTP_HOST = "smtp.yandex.ru"
SMTP_PORT = 465
SMTP_USER = "s9308852555@yandex.ru"
TO_EMAIL  = "s9308852555@yandex.ru"
SCHEMA    = "t_p87775074_welding_suit_factory"

ORANGE = "F57C00"
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
LABEL_FONT  = Font(name="Arial", color="666666", size=10)
VALUE_FONT  = Font(name="Arial", bold=True, size=11)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD")
)


def cors():
    return {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "POST, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type",
    }


def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def build_contact_excel(org, contact, phone, email, message):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка на КП"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45

    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "СПЕЦНАЗ ФАБРИКА — Заявка на КП"
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws["B1"].fill = HEADER_FILL

    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", color="999999", size=9)
    ws["A2"].alignment = Alignment(horizontal="right")

    fields = [
        ("Организация", org),
        ("Контактное лицо", contact),
        ("Телефон", phone),
        ("E-mail", email or "—"),
        ("Что требуется", message),
    ]
    for i, (label, value) in enumerate(fields, start=4):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=1).border = THIN_BORDER
        v = ws.cell(row=i, column=2, value=value)
        v.font = VALUE_FONT
        v.border = THIN_BORDER
        if label == "Что требуется":
            v.alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_order_excel(org, contact, phone, email, with_logo, subtotal, volume_discount, total, groups):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказ"

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14

    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "СПЕЦНАЗ ФАБРИКА — Заказ из калькулятора"
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    for c in ["B1", "C1", "D1", "E1", "F1", "G1", "H1"]:
        ws[c].fill = HEADER_FILL

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", color="999999", size=9)
    ws["A2"].alignment = Alignment(horizontal="right")

    info = [
        ("Организация", org),
        ("Контактное лицо", contact),
        ("Телефон", phone),
        ("E-mail", email or "—"),
    ]
    for i, (label, value) in enumerate(info, start=4):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        v = ws.cell(row=i, column=2, value=value)
        v.font = VALUE_FONT
        v.border = THIN_BORDER

    row = len(info) + 5
    conditions = [("Нанесение логотипа", "Да (+15%)" if with_logo else "Нет")]
    if volume_discount > 0:
        conditions.append(("Скидка за объём", f"{round(volume_discount * 100)}%"))
    for label, value in conditions:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        v = ws.cell(row=row, column=2, value=value)
        v.font = VALUE_FONT
        v.border = THIN_BORDER
        row += 1

    row += 1
    all_items = []
    for g in groups:
        all_items.extend(g.get("items", []))
    has_savings = any(item.get("saving", 0) > 0 for item in all_items)
    NCOLS = 8 if has_savings else 7

    for g in groups:
        g_payment = g.get("payment", "")
        g_desc = g.get("paymentDesc", "")
        g_total = g.get("total", 0)
        g_items = g.get("items", [])
        pay_label = f"{g_payment} ({g_desc})" if g_desc else g_payment

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        gh = ws.cell(row=row, column=1, value=pay_label)
        gh.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        gh.fill = PatternFill(start_color="555555", end_color="555555", fill_type="solid")
        gh.alignment = Alignment(horizontal="center")
        for c2 in range(2, NCOLS + 1):
            ws.cell(row=row, column=c2).fill = PatternFill(start_color="555555", end_color="555555", fill_type="solid")
        row += 1

        headers = ["Артикул", "GTIN", "Размер", "Кол-во", "Цена без скидки", "Цена/шт", "Сумма, ₽"]
        if has_savings:
            headers.append("Экономия, ₽")
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        row += 1

        for item in g_items:
            saving = item.get("saving", 0)
            ws.cell(row=row, column=1, value=item.get("product", "")).font = Font(name="Arial", size=10)
            gtin_val = item.get("gtin", "")
            gc = ws.cell(row=row, column=2, value=gtin_val)
            gc.font = Font(name="Arial", size=10)
            gc.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=item.get("size", "")).font = Font(name="Arial", size=10)
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=item.get("qty", 0)).font = Font(name="Arial", size=10)
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            uf = ws.cell(row=row, column=5, value=item.get("unitPriceFull", 0))
            uf.font = Font(name="Arial", size=10, color="999999")
            uf.alignment = Alignment(horizontal="right")
            uf.number_format = '#,##0'
            up = ws.cell(row=row, column=6, value=item.get("unitPrice", 0))
            up.font = Font(name="Arial", bold=True, size=10)
            up.alignment = Alignment(horizontal="right")
            up.number_format = '#,##0'
            lt = ws.cell(row=row, column=7, value=item.get("lineTotal", 0))
            lt.font = Font(name="Arial", bold=True, size=10)
            lt.alignment = Alignment(horizontal="right")
            lt.number_format = '#,##0'
            if has_savings:
                sv = ws.cell(row=row, column=8, value=f"-{saving:,}" if saving > 0 else "—")
                sv.font = Font(name="Arial", size=10, color="4CAF50" if saving > 0 else "999999")
                sv.alignment = Alignment(horizontal="right")
            for col in range(1, NCOLS + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS - 1)
        ws.cell(row=row, column=1, value=f"Подитог ({pay_label}):").font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        gt = ws.cell(row=row, column=NCOLS, value=g_total)
        gt.font = Font(name="Arial", bold=True, size=11, color=ORANGE)
        gt.alignment = Alignment(horizontal="right")
        gt.number_format = '#,##0'
        row += 2

    total_saving = sum(item.get("saving", 0) for item in all_items)
    if total_saving > 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS - 1)
        ws.cell(row=row, column=1, value="Общая экономия:").font = Font(name="Arial", bold=True, size=11, color="4CAF50")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        ts = ws.cell(row=row, column=NCOLS, value=total_saving)
        ts.font = Font(name="Arial", bold=True, size=11, color="4CAF50")
        ts.alignment = Alignment(horizontal="right")
        ts.number_format = '-#,##0'
        row += 1

    if subtotal != total and volume_discount > 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS - 1)
        ws.cell(row=row, column=1, value="Сумма без скидки:").font = Font(name="Arial", size=11, color="666666")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        st = ws.cell(row=row, column=NCOLS, value=subtotal)
        st.font = Font(name="Arial", size=11, color="666666")
        st.alignment = Alignment(horizontal="right")
        st.number_format = '#,##0'
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS - 1)
        ws.cell(row=row, column=1, value=f"Скидка за объём ({round(volume_discount * 100)}%):").font = Font(name="Arial", size=11, color="4CAF50")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        dc = ws.cell(row=row, column=NCOLS, value=subtotal - total)
        dc.font = Font(name="Arial", size=11, color="4CAF50")
        dc.alignment = Alignment(horizontal="right")
        dc.number_format = '-#,##0'
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS - 1)
    ws.cell(row=row, column=1, value="ИТОГО К ОПЛАТЕ:").font = Font(name="Arial", bold=True, size=13)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    t = ws.cell(row=row, column=NCOLS, value=total)
    t.font = Font(name="Arial", bold=True, size=13, color=ORANGE)
    t.alignment = Alignment(horizontal="right")
    t.number_format = '#,##0'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_email(subject, html, to=None, excel_bytes=None, excel_filename="Заявка.xlsx"):
    pwd = os.environ.get("SMTP_PASSWORD", "")
    recipients = to if to else [TO_EMAIL]
    try:
        msg = MIMEMultipart("mixed")
        msg["Subject"] = subject
        msg["From"]    = SMTP_USER
        msg["To"]      = ", ".join(recipients)

        html_part = MIMEMultipart("alternative")
        html_part.attach(MIMEText(html, "html", "utf-8"))
        msg.attach(html_part)

        if excel_bytes:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(excel_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment", filename=excel_filename)
            msg.attach(part)

        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
            server.login(SMTP_USER, pwd)
            server.sendmail(SMTP_USER, recipients, msg.as_string())
        print(f"SMTP: email sent OK to {recipients}")
        return True
    except Exception as e:
        print(f"SMTP error: {e}")
        return False


def build_client_html(contact, with_logo, subtotal, volume_discount, total, groups):
    """HTML-письмо клиенту — табличный формат как Excel."""
    name = contact if contact and contact != "—" else "Уважаемый клиент"
    logo_label = "Да (+15%)" if with_logo else "Нет"
    discount_pct = round(volume_discount * 100) if volume_discount else 0
    td = 'style="padding:5px 8px;border:1px solid #ddd;font-size:12px'

    cli_all_items = []
    for gg in groups:
        cli_all_items.extend(gg.get("items", []))
    cli_has_savings = any(i.get("saving", 0) > 0 for i in cli_all_items)
    cli_total_saving = sum(i.get("saving", 0) for i in cli_all_items)
    cli_ncols = 7 if cli_has_savings else 6

    groups_html = ""
    for g in groups:
        g_payment = g.get("payment", "")
        g_desc = g.get("paymentDesc", "")
        g_total = g.get("total", 0)
        g_items = g.get("items", [])
        pay_label = f"{g_payment} ({g_desc})" if g_desc else g_payment

        rows = ""
        for item in g_items:
            saving = item.get('saving', 0)
            saving_td = f'<td {td};text-align:right;color:#4CAF50">−{saving:,}</td>' if cli_has_savings and saving > 0 else (f'<td {td};text-align:right;color:#999">—</td>' if cli_has_savings else "")
            rows += f"""
      <tr>
        <td {td}">{item.get('product','')}</td>
        <td {td};text-align:center">{item.get('gtin','')}</td>
        <td {td};text-align:center">{item.get('size','')}</td>
        <td {td};text-align:center">{item.get('qty','')}</td>
        <td {td};text-align:right;font-weight:bold">{item.get('unitPrice',0):,}</td>
        <td {td};text-align:right;font-weight:bold">{item.get('lineTotal',0):,}</td>
        {saving_td}
      </tr>"""

        saving_th = f'<th {td};color:#fff;text-align:right">Экономия</th>' if cli_has_savings else ""
        groups_html += f"""
    <table style="width:100%;border-collapse:collapse;margin-bottom:4px">
      <tr><td colspan="{cli_ncols}" style="background:#555;color:#fff;padding:6px 8px;font-size:12px;font-weight:bold;letter-spacing:1px">{pay_label}</td></tr>
      <tr style="background:#f57c00">
        <th {td};color:#fff">Артикул</th>
        <th {td};color:#fff;text-align:center">GTIN</th>
        <th {td};color:#fff;text-align:center">Размер</th>
        <th {td};color:#fff;text-align:center">Кол-во</th>
        <th {td};color:#fff;text-align:right">Цена/шт</th>
        <th {td};color:#fff;text-align:right">Сумма</th>
        {saving_th}
      </tr>
      {rows}
      <tr style="background:#f5f5f5">
        <td colspan="{cli_ncols - 1}" {td};text-align:right;color:#666">Подитог:</td>
        <td {td};text-align:right;font-weight:bold;color:#f57c00;font-size:13px">{g_total:,} ₽</td>
      </tr>
    </table>"""

    footer_rows = ""
    if cli_total_saving > 0:
        footer_rows += f"""
      <tr><td colspan="{cli_ncols - 1}" {td};text-align:right;color:#4CAF50;font-weight:bold;border:none">Общая экономия:</td><td {td};text-align:right;color:#4CAF50;font-weight:bold;border:none">−{cli_total_saving:,} ₽</td></tr>"""
    if subtotal != total and discount_pct > 0:
        footer_rows += f"""
      <tr><td colspan="{cli_ncols - 1}" {td};text-align:right;color:#666;border:none">Сумма без скидки:</td><td {td};text-align:right;color:#666;border:none">{subtotal:,} ₽</td></tr>
      <tr><td colspan="{cli_ncols - 1}" {td};text-align:right;color:#4CAF50;border:none">Скидка за объём ({discount_pct}%):</td><td {td};text-align:right;color:#4CAF50;border:none">−{subtotal - total:,} ₽</td></tr>"""

    ccr = cli_ncols - 1
    return f"""
<div style="font-family:Arial,sans-serif;max-width:750px;margin:0 auto">
  <table style="width:100%;border-collapse:collapse">
    <tr><td colspan="{cli_ncols}" style="background:#f57c00;padding:12px 16px">
      <span style="color:#fff;font-size:18px;font-weight:bold;letter-spacing:2px">СПЕЦНАЗ ФАБРИКА</span>
    </td></tr>
    <tr><td colspan="{cli_ncols}" style="padding:12px 8px;font-size:14px;color:#222">
      {name}, благодарим Вас за обращение!<br>
      Ваша заявка принята. Менеджер свяжется с Вами в ближайшее время.
    </td></tr>
    <tr><td {td};color:#666;width:160px">Нанесение логотипа</td><td colspan="{ccr}" {td}">{logo_label}</td></tr>
    {"<tr><td " + td + ";color:#666" + '">Скидка за объём</td><td colspan="' + str(ccr) + '" ' + td + ";color:#4CAF50;font-weight:bold" + '">' + str(discount_pct) + "%</td></tr>" if discount_pct > 0 else ""}
  </table>
  <div style="height:12px"></div>
  {groups_html}
  <table style="width:100%;border-collapse:collapse">
    {footer_rows}
    <tr style="background:#fff3e0"><td colspan="{ccr}" {td};text-align:right;font-weight:bold;font-size:14px;border:2px solid #f57c00">ИТОГО:</td><td {td};text-align:right;font-weight:bold;font-size:16px;color:#f57c00;border:2px solid #f57c00">{total:,} ₽</td></tr>
  </table>
  <table style="width:100%;border-collapse:collapse;margin-top:16px">
    <tr><td {td};color:#666;border:none;width:80px">Телефон:</td><td {td};border:none"><a href="tel:+79308852555" style="color:#f57c00;text-decoration:none;font-weight:bold">8-930-885-25-55</a></td></tr>
    <tr><td {td};color:#666;border:none">E-mail:</td><td {td};border:none"><a href="mailto:s9308852555@yandex.ru" style="color:#f57c00;text-decoration:none">s9308852555@yandex.ru</a></td></tr>
    <tr><td {td};color:#666;border:none">Сайт:</td><td {td};border:none"><a href="https://спецназфабрика.рф" style="color:#f57c00;text-decoration:none">спецназфабрика.рф</a></td></tr>
  </table>
  <div style="background:#eee;padding:8px 16px;font-size:11px;color:#999;margin-top:8px">
    С уважением, команда СПЕЦНАЗ ФАБРИКА · <a href="https://спецназфабрика.рф" style="color:#f57c00">спецназфабрика.рф</a>
  </div>
</div>"""


def save_lead(org, contact, phone, email, message, kind, order_json, order_total):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(
        f"""INSERT INTO {SCHEMA}.leads (org, contact, phone, email, message, kind, order_json, order_total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
        (org, contact, phone, email, message, kind, order_json, order_total)
    )
    conn.commit()
    conn.close()


def handler(event: dict, context) -> dict:
    """Принимает заявку с сайта спецназфабрика.рф, сохраняет контакт в БД и отправляет письмо с Excel-файлом менеджеру."""

    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": cors(), "body": ""}

    body    = json.loads(event.get("body") or "{}")
    kind    = body.get("kind", "contact")
    org     = body.get("org", "—")
    contact = body.get("contact", "—")
    phone   = body.get("phone", "—")
    email   = body.get("email", "")

    if kind == "contact":
        message = body.get("message", "—")
        save_lead(org, contact, phone, email, message, "contact", "", 0)

        excel = build_contact_excel(org, contact, phone, email, message)

        html = f"""
<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
  <div style="background:#f57c00;padding:16px 24px">
    <h1 style="color:#fff;margin:0;font-size:20px;letter-spacing:2px">СПЕЦНАЗ ФАБРИКА — новая заявка</h1>
  </div>
  <div style="background:#f9f9f9;padding:24px;border:1px solid #eee">
    <table style="width:100%;border-collapse:collapse">
      <tr><td style="padding:8px 0;color:#666;width:180px">Организация</td><td style="padding:8px 0;font-weight:bold;color:#222">{org}</td></tr>
      <tr><td style="padding:8px 0;color:#666">Контактное лицо</td><td style="padding:8px 0;font-weight:bold;color:#222">{contact}</td></tr>
      <tr><td style="padding:8px 0;color:#666">Телефон</td><td style="padding:8px 0;font-weight:bold;color:#f57c00;font-size:16px">{phone}</td></tr>
      <tr><td style="padding:8px 0;color:#666">E-mail</td><td style="padding:8px 0;color:#222">{email or '—'}</td></tr>
      <tr><td style="padding:8px 0;color:#666;vertical-align:top">Что требуется</td><td style="padding:8px 0;color:#222">{message}</td></tr>
    </table>
  </div>
  <div style="background:#eee;padding:12px 24px;font-size:12px;color:#999">
    Заявка отправлена с сайта спецназфабрика.рф · Excel-файл во вложении
  </div>
</div>"""
        fname = f"Заявка_КП_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        send_email("Новая заявка на КП — спецназфабрика.рф", html, excel_bytes=excel, excel_filename=fname)

        if email and "@" in email:
            name = contact if contact and contact != "—" else "Уважаемый клиент"
            client_html = f"""
<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
  <div style="background:#f57c00;padding:16px 24px">
    <h1 style="color:#fff;margin:0;font-size:20px;letter-spacing:2px">СПЕЦНАЗ ФАБРИКА</h1>
  </div>
  <div style="background:#f9f9f9;padding:24px;border:1px solid #eee">
    <p style="color:#222;font-size:15px;margin:0 0 16px">
      {name}, благодарим Вас за обращение!<br>
      Ваша заявка на коммерческое предложение принята. Наш менеджер свяжется с Вами в течение 2 часов.
    </p>
    <p style="color:#222;font-size:14px;margin:16px 0 0;line-height:1.6">Если у Вас возникнут вопросы, мы всегда на связи:</p>
    <table style="width:100%;border-collapse:collapse;margin-top:8px">
      <tr><td style="padding:4px 0;color:#666;font-size:13px;width:80px">Телефон:</td><td style="padding:4px 0;font-size:13px"><a href="tel:+79308852555" style="color:#f57c00;text-decoration:none;font-weight:bold">8-930-885-25-55</a></td></tr>
      <tr><td style="padding:4px 0;color:#666;font-size:13px">E-mail:</td><td style="padding:4px 0;font-size:13px"><a href="mailto:s9308852555@yandex.ru" style="color:#f57c00;text-decoration:none">s9308852555@yandex.ru</a></td></tr>
      <tr><td style="padding:4px 0;color:#666;font-size:13px">Сайт:</td><td style="padding:4px 0;font-size:13px"><a href="https://спецназфабрика.рф" style="color:#f57c00;text-decoration:none">спецназфабрика.рф</a></td></tr>
    </table>
  </div>
  <div style="background:#eee;padding:12px 24px;font-size:12px;color:#999">
    С уважением, команда СПЕЦНАЗ ФАБРИКА · <a href="https://спецназфабрика.рф" style="color:#f57c00">спецназфабрика.рф</a>
  </div>
</div>"""
            send_email("Ваша заявка принята — СПЕЦНАЗ ФАБРИКА", client_html, to=[email])

    elif kind == "order":
        with_logo       = body.get("withLogo", False)
        subtotal        = body.get("subtotal", 0)
        volume_discount = body.get("volumeDiscount", 0)
        total           = body.get("total", 0)
        groups          = body.get("groups", [])

        all_items = []
        for g in groups:
            all_items.extend(g.get("items", []))
        save_lead(org, contact, phone, email, "", "order", json.dumps({"groups": groups}, ensure_ascii=False), total)

        excel = build_order_excel(org, contact, phone, email, with_logo, subtotal, volume_discount, total, groups)

        logo_label = "Да (+15%)" if with_logo else "Нет"
        discount_pct = round(volume_discount * 100) if volume_discount else 0

        td = 'style="padding:5px 8px;border:1px solid #ddd;font-size:12px'
        mgr_has_savings = any(item.get("saving", 0) > 0 for item in all_items)
        mgr_total_saving = sum(item.get("saving", 0) for item in all_items)
        ncols_html = 8 if mgr_has_savings else 7

        groups_html = ""
        for g in groups:
            g_payment = g.get("payment", "")
            g_desc = g.get("paymentDesc", "")
            g_total = g.get("total", 0)
            g_items = g.get("items", [])
            pay_label = f"{g_payment} ({g_desc})" if g_desc else g_payment

            rows = ""
            for item in g_items:
                saving = item.get('saving', 0)
                saving_td = f'<td {td};text-align:right;color:#4CAF50">−{saving:,}</td>' if mgr_has_savings and saving > 0 else (f'<td {td};text-align:right;color:#999">—</td>' if mgr_has_savings else "")
                rows += f"""
      <tr>
        <td {td}">{item.get('product','')}</td>
        <td {td};text-align:center">{item.get('gtin','')}</td>
        <td {td};text-align:center">{item.get('size','')}</td>
        <td {td};text-align:center">{item.get('qty','')}</td>
        <td {td};text-align:right;color:#999">{item.get('unitPriceFull', item.get('unitPrice',0)):,}</td>
        <td {td};text-align:right;font-weight:bold">{item.get('unitPrice',0):,}</td>
        <td {td};text-align:right;font-weight:bold">{item.get('lineTotal',0):,}</td>
        {saving_td}
      </tr>"""

            saving_th = f'<th {td};color:#fff;text-align:right">Экономия</th>' if mgr_has_savings else ""
            groups_html += f"""
    <table style="width:100%;border-collapse:collapse;margin-bottom:4px">
      <tr><td colspan="{ncols_html}" style="background:#555;color:#fff;padding:6px 8px;font-size:12px;font-weight:bold;letter-spacing:1px">{pay_label}</td></tr>
      <tr style="background:#f57c00">
        <th {td};color:#fff">Артикул</th>
        <th {td};color:#fff;text-align:center">GTIN</th>
        <th {td};color:#fff;text-align:center">Размер</th>
        <th {td};color:#fff;text-align:center">Кол-во</th>
        <th {td};color:#fff;text-align:right">Цена до скидки</th>
        <th {td};color:#fff;text-align:right">Цена/шт</th>
        <th {td};color:#fff;text-align:right">Сумма</th>
        {saving_th}
      </tr>
      {rows}
      <tr style="background:#f5f5f5">
        <td colspan="{ncols_html - 1}" {td};text-align:right;color:#666">Подитог ({pay_label}):</td>
        <td {td};text-align:right;font-weight:bold;color:#f57c00;font-size:13px">{g_total:,} ₽</td>
      </tr>
    </table>"""

        footer_rows = ""
        if mgr_total_saving > 0:
            footer_rows += f"""
      <tr><td colspan="{ncols_html - 1}" {td};text-align:right;color:#4CAF50;font-weight:bold;border:none">Общая экономия:</td><td {td};text-align:right;color:#4CAF50;font-weight:bold;border:none">−{mgr_total_saving:,} ₽</td></tr>"""
        if subtotal != total and discount_pct > 0:
            footer_rows += f"""
      <tr><td colspan="{ncols_html - 1}" {td};text-align:right;color:#666;border:none">Сумма без скидки:</td><td {td};text-align:right;color:#666;border:none">{subtotal:,} ₽</td></tr>
      <tr><td colspan="{ncols_html - 1}" {td};text-align:right;color:#4CAF50;border:none">Скидка за объём ({discount_pct}%):</td><td {td};text-align:right;color:#4CAF50;border:none">−{subtotal - total:,} ₽</td></tr>"""

        cs = ncols_html
        csr = cs - 1
        html = f"""
<div style="font-family:Arial,sans-serif;max-width:750px;margin:0 auto">
  <table style="width:100%;border-collapse:collapse">
    <tr><td colspan="{cs}" style="background:#f57c00;padding:12px 16px">
      <span style="color:#fff;font-size:18px;font-weight:bold;letter-spacing:2px">СПЕЦНАЗ ФАБРИКА — Заказ из калькулятора</span>
    </td></tr>
    <tr><td colspan="{cs}" style="padding:4px 8px;text-align:right;color:#999;font-size:11px">Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}</td></tr>
    <tr><td {td};color:#666;width:160px">Организация</td><td colspan="{csr}" {td};font-weight:bold">{org}</td></tr>
    <tr><td {td};color:#666">Контактное лицо</td><td colspan="{csr}" {td};font-weight:bold">{contact}</td></tr>
    <tr><td {td};color:#666">Телефон</td><td colspan="{csr}" {td};font-weight:bold;color:#f57c00">{phone}</td></tr>
    <tr><td {td};color:#666">E-mail</td><td colspan="{csr}" {td}">{email or '—'}</td></tr>
    <tr><td {td};color:#666">Нанесение логотипа</td><td colspan="{csr}" {td}">{logo_label}</td></tr>
    {"<tr><td " + td + ";color:#666" + '">Скидка за объём</td><td colspan="' + str(csr) + '" ' + td + ";color:#4CAF50;font-weight:bold" + '">' + str(discount_pct) + "%</td></tr>" if discount_pct > 0 else ""}
  </table>
  <div style="height:12px"></div>
  {groups_html}
  <table style="width:100%;border-collapse:collapse">
    {footer_rows}
    <tr style="background:#fff3e0"><td colspan="{csr}" {td};text-align:right;font-weight:bold;font-size:14px;border:2px solid #f57c00">ИТОГО К ОПЛАТЕ:</td><td {td};text-align:right;font-weight:bold;font-size:16px;color:#f57c00;border:2px solid #f57c00">{total:,} ₽</td></tr>
  </table>
  <div style="background:#eee;padding:8px 16px;font-size:11px;color:#999;margin-top:8px">
    Заявка отправлена с сайта спецназфабрика.рф · Excel-файл во вложении
  </div>
</div>"""
        fname = f"Заказ_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        send_email(f"Заказ на {total:,} ₽ — спецназфабрика.рф", html, excel_bytes=excel, excel_filename=fname)

        if email and "@" in email:
            client_html = build_client_html(contact, with_logo, subtotal, volume_discount, total, groups)
            send_email(f"Ваш заказ на {total:,} ₽ — СПЕЦНАЗ ФАБРИКА", client_html, to=[email], excel_bytes=excel, excel_filename=fname)

    else:
        return {"statusCode": 400, "headers": {**cors(), "Content-Type": "application/json"}, "body": json.dumps({"error": "unknown kind"})}

    return {
        "statusCode": 200,
        "headers": {**cors(), "Content-Type": "application/json"},
        "body": json.dumps({"ok": True}),
    }